# reportes/views.py — COMPLETO CON PERÍODOS Y FILTRO POR CATEGORÍA
#
# Novedad vs versión anterior:
# - Filtro por período: diario, semanal, mensual, semestral, anual
# - Filtro por categoría (ej: solo ver "Pyme")
# - Cálculo de promedio diario, proyección mensual y anual por categoría
# - La vista construye el rango de fechas según el período elegido

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Sum, Q
from gastos.models import Gasto, Ingreso, Categoria
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import json
from datetime import date, timedelta
from decimal import Decimal


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

MESES_LISTA = [
    (1,'Enero'),(2,'Febrero'),(3,'Marzo'),(4,'Abril'),
    (5,'Mayo'),(6,'Junio'),(7,'Julio'),(8,'Agosto'),
    (9,'Septiembre'),(10,'Octubre'),(11,'Noviembre'),(12,'Diciembre'),
]

MESES_CORTO = ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']


def _rango_fechas(periodo, año, mes):
    """
    Devuelve (fecha_inicio, fecha_fin, titulo, dias) según el período elegido.
    
    ¿Para qué sirve esto?
    En vez de filtrar siempre por mes completo, ahora podemos filtrar
    por cualquier rango: una semana, un día, un semestre, un año.
    Todos los cálculos posteriores usan este rango.
    """
    hoy = timezone.now().date()

    if periodo == 'diario':
        # El día actual
        inicio = date(año, mes, hoy.day) if (año == hoy.year and mes == hoy.month) else date(año, mes, 1)
        fin = inicio
        titulo = f"Día {inicio.strftime('%d/%m/%Y')}"
        dias = 1

    elif periodo == 'semanal':
        # La semana actual (lunes a domingo)
        lunes = hoy - timedelta(days=hoy.weekday())
        inicio = lunes
        fin = lunes + timedelta(days=6)
        titulo = f"Semana {inicio.strftime('%d/%m')} – {fin.strftime('%d/%m/%Y')}"
        dias = 7

    elif periodo == 'semestral':
        # Semestre del año según el mes elegido
        if mes <= 6:
            inicio = date(año, 1, 1)
            fin = date(año, 6, 30)
            titulo = f"1° Semestre {año}"
        else:
            inicio = date(año, 7, 1)
            fin = date(año, 12, 31)
            titulo = f"2° Semestre {año}"
        dias = (fin - inicio).days + 1

    elif periodo == 'anual':
        inicio = date(año, 1, 1)
        fin = date(año, 12, 31)
        titulo = f"Año {año}"
        dias = 365

    else:  # mensual (default)
        import calendar
        ultimo_dia = calendar.monthrange(año, mes)[1]
        inicio = date(año, mes, 1)
        fin = date(año, mes, ultimo_dia)
        titulo = f"{dict(MESES_LISTA)[mes]} {año}"
        dias = ultimo_dia

    return inicio, fin, titulo, dias


def _datos_categorias_json(gastos):
    """Prepara JSON para el gráfico de torta de categorías."""
    por_cat = (
        gastos.values('categoria__nombre', 'categoria__icono')
        .annotate(subtotal=Sum('monto'))
        .order_by('-subtotal')
    )
    labels, values = [], []
    for item in por_cat:
        nombre = item['categoria__nombre'] or 'Sin categoría'
        icono = item['categoria__icono'] or ''
        labels.append(f"{icono} {nombre}".strip())
        values.append(float(item['subtotal'] or 0))
    return json.dumps({'labels': labels, 'values': values})


def _resumen_por_categoria(gastos, dias, total_egresos):
    """
    Para cada categoría calcula:
    - total en el período
    - promedio diario
    - proyección mensual (× 30 días)
    - proyección anual (× 365 días)
    - porcentaje del total de egresos
    
    Esto responde exactamente tu pregunta: ¿cuánto gasto en Pyme
    por día, por mes, por año?
    """
    por_cat = (
        gastos.values('categoria__nombre', 'categoria__icono')
        .annotate(subtotal=Sum('monto'))
        .order_by('-subtotal')
    )
    resultado = []
    for item in por_cat:
        total = float(item['subtotal'] or 0)
        if total == 0 or dias == 0:
            continue
        por_dia = total / dias
        resultado.append({
            'nombre': f"{item['categoria__icono'] or ''} {item['categoria__nombre'] or 'Sin categoría'}".strip(),
            'total': total,
            'por_dia': round(por_dia),
            'proyeccion_mensual': round(por_dia * 30),
            'proyeccion_anual': round(por_dia * 365),
            'porcentaje': round((total / total_egresos * 100)) if total_egresos > 0 else 0,
        })
    return resultado


def _evolucion(usuario, periodo, año, mes):
    """
    Genera los datos para el gráfico de barras de evolución.
    El período determina qué se compara:
    - mensual/diario/semanal → últimos 6 meses
    - semestral → últimos 4 semestres
    - anual → últimos 5 años
    """
    labels, egresos_vals, ingresos_vals = [], [], []

    if periodo == 'anual':
        titulo = "Evolución últimos 5 años"
        for a in range(año - 4, año + 1):
            labels.append(str(a))
            eg = Gasto.objects.filter(usuario=usuario, fecha__year=a).aggregate(t=Sum('monto'))['t'] or 0
            ing = Ingreso.objects.filter(usuario=usuario, fecha__year=a).aggregate(t=Sum('monto'))['t'] or 0
            egresos_vals.append(float(eg))
            ingresos_vals.append(float(ing))

    elif periodo == 'semestral':
        titulo = "Evolución últimos 4 semestres"
        # Generar los últimos 4 semestres
        semestres = []
        a, s = año, (1 if mes <= 6 else 2)
        for _ in range(4):
            semestres.insert(0, (a, s))
            s -= 1
            if s == 0:
                s = 2
                a -= 1
        for a_s, s_s in semestres:
            if s_s == 1:
                inicio, fin = date(a_s, 1, 1), date(a_s, 6, 30)
                labels.append(f"1S {a_s}")
            else:
                inicio, fin = date(a_s, 7, 1), date(a_s, 12, 31)
                labels.append(f"2S {a_s}")
            eg = Gasto.objects.filter(usuario=usuario, fecha__range=(inicio, fin)).aggregate(t=Sum('monto'))['t'] or 0
            ing = Ingreso.objects.filter(usuario=usuario, fecha__range=(inicio, fin)).aggregate(t=Sum('monto'))['t'] or 0
            egresos_vals.append(float(eg))
            ingresos_vals.append(float(ing))

    else:  # mensual, semanal, diario → últimos 6 meses
        titulo = "Evolución últimos 6 meses"
        m, a = mes, año
        meses = []
        for _ in range(6):
            meses.insert(0, (a, m))
            m -= 1
            if m == 0:
                m, a = 12, a - 1
        for a_m, m_m in meses:
            labels.append(f"{MESES_CORTO[m_m-1]} {str(a_m)[2:]}")
            eg = Gasto.objects.filter(usuario=usuario, fecha__year=a_m, fecha__month=m_m).aggregate(t=Sum('monto'))['t'] or 0
            ing = Ingreso.objects.filter(usuario=usuario, fecha__year=a_m, fecha__month=m_m).aggregate(t=Sum('monto'))['t'] or 0
            egresos_vals.append(float(eg))
            ingresos_vals.append(float(ing))

    return (
        json.dumps(labels),
        json.dumps(egresos_vals),
        json.dumps(ingresos_vals),
        titulo,
    )


# ─────────────────────────────────────────────
# VISTA PRINCIPAL
# ─────────────────────────────────────────────

@login_required
def estadisticas(request):
    hoy = timezone.now().date()
    año = int(request.GET.get('año', hoy.year))
    mes = int(request.GET.get('mes', hoy.month))
    periodo = request.GET.get('periodo', 'mensual')
    categoria_filtro = request.GET.get('categoria_filtro', '')

    # Rango de fechas según período
    fecha_inicio, fecha_fin, titulo_periodo, dias = _rango_fechas(periodo, año, mes)

    # Gastos del período
    gastos_qs = Gasto.objects.filter(
        usuario=request.user,
        fecha__range=(fecha_inicio, fecha_fin),
    ).select_related('categoria').order_by('estado', '-fecha')

    # Filtro por categoría (opcional)
    categoria_nombre = ''
    if categoria_filtro:
        gastos_qs = gastos_qs.filter(categoria__pk=categoria_filtro)
        try:
            cat = Categoria.objects.get(pk=categoria_filtro)
            categoria_nombre = str(cat)
        except Categoria.DoesNotExist:
            categoria_filtro = ''

    # Ingresos (sin filtro de categoría — ingresos no tienen categoría)
    ingresos_qs = Ingreso.objects.filter(
        usuario=request.user,
        fecha__range=(fecha_inicio, fecha_fin),
    )

    # Totales
    total_egresos = gastos_qs.aggregate(t=Sum('monto'))['t'] or 0
    total_pendiente = gastos_qs.filter(estado='pendiente').aggregate(t=Sum('monto'))['t'] or 0
    total_ingresos = ingresos_qs.aggregate(t=Sum('monto'))['t'] or 0
    balance = total_ingresos - total_egresos

    # Promedios diarios
    promedio_diario_egresos = round(float(total_egresos) / dias) if dias > 0 and total_egresos else 0
    promedio_diario_ingresos = round(float(total_ingresos) / dias) if dias > 0 and total_ingresos else 0

    # Datos para gráficos
    por_categoria = gastos_qs.values('categoria__nombre').annotate(subtotal=Sum('monto')).filter(subtotal__gt=0)
    ingresos_por_tipo = ingresos_qs.values('tipo').annotate(subtotal=Sum('monto'))

    datos_categorias_json = _datos_categorias_json(gastos_qs)

    TIPOS_INGRESO = {'sueldo': 'Sueldo', 'extra': 'Extra', 'pyme': 'Pyme', 'otro': 'Otro'}
    datos_ingresos_json = json.dumps({
        'labels': [TIPOS_INGRESO.get(i['tipo'], i['tipo']) for i in ingresos_por_tipo],
        'values': [float(i['subtotal'] or 0) for i in ingresos_por_tipo],
    })

    evolucion_labels, evolucion_egresos, evolucion_ingresos, titulo_evolucion = _evolucion(
        request.user, periodo, año, mes
    )

    # Resumen detallado por categoría (con proyecciones)
    resumen_categorias = _resumen_por_categoria(gastos_qs, dias, float(total_egresos))

    return render(request, 'gastos/estadisticas.html', {
        # Filtros
        'periodo': periodo,
        'mes': mes,
        'año': año,
        'categoria_filtro': categoria_filtro,
        'categoria_nombre': categoria_nombre,
        'categorias': Categoria.objects.all(),
        'meses_lista': MESES_LISTA,

        # Títulos
        'titulo_periodo': titulo_periodo,
        'titulo_evolucion': titulo_evolucion,

        # Datos
        'gastos': gastos_qs,
        'ingresos': ingresos_qs,
        'total_egresos': total_egresos,
        'total_pendiente': total_pendiente,
        'total_ingresos': total_ingresos,
        'balance': balance,
        'promedio_diario_egresos': promedio_diario_egresos,
        'promedio_diario_ingresos': promedio_diario_ingresos,
        'por_categoria': por_categoria,
        'ingresos_por_tipo': ingresos_por_tipo,
        'resumen_categorias': resumen_categorias,

        # JSON para Chart.js
        'datos_categorias_json': datos_categorias_json,
        'datos_ingresos_json': datos_ingresos_json,
        'evolucion_labels': evolucion_labels,
        'evolucion_egresos': evolucion_egresos,
        'evolucion_ingresos': evolucion_ingresos,
    })


# ─────────────────────────────────────────────
# EXPORT EXCEL
# ─────────────────────────────────────────────

@login_required
def exportar_excel(request):
    hoy = timezone.now().date()
    gastos = Gasto.objects.filter(
        usuario=request.user,
        fecha__year=hoy.year,
        fecha__month=hoy.month,
    ).select_related('categoria')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Gastos {hoy.month}-{hoy.year}"

    enc_font = Font(bold=True, color="FFFFFF")
    enc_fill = PatternFill("solid", fgColor="1a1a2e")
    encabezados = ['Descripción','Monto','Categoría','Fecha','Vencimiento','Prioridad','Estado']

    for col, titulo in enumerate(encabezados, 1):
        c = ws.cell(row=1, column=col, value=titulo)
        c.font = enc_font
        c.fill = enc_fill
        c.alignment = Alignment(horizontal='center')

    for row, g in enumerate(gastos, 2):
        ws.cell(row=row, column=1, value=g.descripcion)
        ws.cell(row=row, column=2, value=float(g.monto) if g.monto else 0)
        ws.cell(row=row, column=3, value=str(g.categoria) if g.categoria else '')
        ws.cell(row=row, column=4, value=str(g.fecha))
        ws.cell(row=row, column=5, value=str(g.fecha_vencimiento) if g.fecha_vencimiento else '')
        ws.cell(row=row, column=6, value=g.get_prioridad_display())
        ws.cell(row=row, column=7, value=g.get_estado_display())

    for col in ws.columns:
        max_length = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 4

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=gastos_{hoy.month}_{hoy.year}.xlsx'
    wb.save(response)
    return response
